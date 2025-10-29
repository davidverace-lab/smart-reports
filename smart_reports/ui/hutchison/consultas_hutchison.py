"""
Panel Consultas Hutchison - Implementación Completa
Panel de consultas con todas las funcionalidades según diseño de capturas
Incluye tabla estilo Excel corporativo y exportación a Excel
"""

import customtkinter as ctk
from tkinter import messagebox
from datetime import datetime
import logging
from smart_reports.config.settings_hutchison import (
    get_color,
    get_font
)

# Configurar logging para debugging de SQL
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)


class ConsultasHutchison(ctk.CTkFrame):
    """
    Panel de consultas completo con:
    - Búsqueda por User ID
    - Consulta por Unidad de Negocio
    - Toggle de Filtros Avanzados
    - Consultas rápidas
    - Tabla tksheet con 9 columnas
    """

    def __init__(self, parent, db=None):
        super().__init__(parent, fg_color=get_color('White'))

        self.db = db
        self.cursor = db.get_cursor() if db else None
        self.sheet = None
        self.advanced_filters_frame = None
        self.show_filters_var = ctk.BooleanVar(value=False)
        self.results_count_label = None
        self.current_data = []  # Almacenar datos actuales para exportar

        self.grid_rowconfigure(0, weight=1)
        self.grid_columnconfigure(0, weight=1)

        self._create_interface()

    def _create_interface(self):
        """Crear interfaz según diseño de capturas"""

        # Scroll frame principal
        scroll_frame = ctk.CTkScrollableFrame(
            self,
            fg_color='transparent'
        )
        scroll_frame.grid(row=0, column=0, sticky='nsew', padx=20, pady=20)

        # ═══════════════════════════════════════════════════
        # TÍTULO
        # ═══════════════════════════════════════════════════
        title_label = ctk.CTkLabel(
            scroll_frame,
            text='Consultas de Usuarios',
            font=('Montserrat', 24, 'bold'),
            text_color=get_color('Sea Blue'),
            anchor='w'
        )
        title_label.pack(fill='x', pady=(0, 20))

        # ═══════════════════════════════════════════════════
        # SECCIÓN 1: BUSCAR POR ID
        # ═══════════════════════════════════════════════════
        self._create_search_by_id(scroll_frame)

        # ═══════════════════════════════════════════════════
        # SECCIÓN 2: BUSCAR POR UNIDAD
        # ═══════════════════════════════════════════════════
        self._create_search_by_unit(scroll_frame)

        # ═══════════════════════════════════════════════════
        # SECCIÓN 3: MOSTRAR FILTROS AVANZADOS (TOGGLE)
        # ═══════════════════════════════════════════════════
        self._create_advanced_filters_toggle(scroll_frame)

        # ═══════════════════════════════════════════════════
        # SECCIÓN 4: CONSULTAS RÁPIDAS
        # ═══════════════════════════════════════════════════
        self._create_quick_queries(scroll_frame)

        # ═══════════════════════════════════════════════════
        # SECCIÓN 5: RESULTADOS (TABLA)
        # ═══════════════════════════════════════════════════
        self._create_results_table(scroll_frame)

    def _create_search_by_id(self, parent):
        """Sección: Buscar Usuario por ID"""
        frame = ctk.CTkFrame(
            parent,
            fg_color='white',
            border_width=1,
            border_color=get_color('Border'),
            corner_radius=10
        )
        frame.pack(fill='x', pady=(0, 15), padx=5)

        inner = ctk.CTkFrame(frame, fg_color='transparent')
        inner.pack(fill='x', padx=20, pady=15)

        # Label
        label = ctk.CTkLabel(
            inner,
            text='🔍  Buscar Usuario por ID:',
            font=('Montserrat', 13, 'bold'),
            text_color=get_color('Sea Blue')
        )
        label.pack(side='left', padx=(0, 15))

        # Entry
        self.id_entry = ctk.CTkEntry(
            inner,
            placeholder_text='Ingrese ID de usuario...',
            width=300,
            height=38,
            font=('Arial', 11),
            border_width=1,
            border_color=get_color('Border')
        )
        self.id_entry.pack(side='left', padx=(0, 10))
        self.id_entry.bind('<Return>', lambda e: self.buscar_por_id())

        # Botón Buscar
        btn_buscar = ctk.CTkButton(
            inner,
            text='Buscar',
            width=120,
            height=38,
            font=('Arial', 11, 'bold'),
            fg_color=get_color('Sky Blue'),
            hover_color=get_color('Sea Blue'),
            text_color='white',
            corner_radius=8,
            command=self.buscar_por_id
        )
        btn_buscar.pack(side='left')

    def _create_search_by_unit(self, parent):
        """Sección: Consultar por Unidad de Negocio"""
        frame = ctk.CTkFrame(
            parent,
            fg_color='white',
            border_width=1,
            border_color=get_color('Border'),
            corner_radius=10
        )
        frame.pack(fill='x', pady=(0, 15), padx=5)

        inner = ctk.CTkFrame(frame, fg_color='transparent')
        inner.pack(fill='x', padx=20, pady=15)

        # Label
        label = ctk.CTkLabel(
            inner,
            text='🏢  Consultar por Unidad de Negocio:',
            font=('Montserrat', 13, 'bold'),
            text_color=get_color('Sea Blue')
        )
        label.pack(side='left', padx=(0, 15))

        # Cargar unidades desde BD
        unidades = self.cargar_unidades()

        # OptionMenu
        self.unit_menu = ctk.CTkOptionMenu(
            inner,
            values=unidades,
            width=250,
            height=38,
            font=('Arial', 11),
            fg_color=get_color('Surface'),
            button_color=get_color('Sky Blue'),
            button_hover_color=get_color('Sea Blue'),
            text_color=get_color('Sea Blue')
        )
        self.unit_menu.pack(side='left', padx=(0, 10))

        # Botón Consultar
        btn_consultar = ctk.CTkButton(
            inner,
            text='Consultar',
            width=120,
            height=38,
            font=('Arial', 11, 'bold'),
            fg_color='#00d4aa',  # Cyan/Turquesa
            hover_color='#00b894',
            text_color='white',
            corner_radius=8,
            command=self.consultar_por_unidad
        )
        btn_consultar.pack(side='left')

    def _create_advanced_filters_toggle(self, parent):
        """Sección: Checkbox para mostrar/ocultar filtros avanzados"""
        # Frame principal para el checkbox
        frame = ctk.CTkFrame(
            parent,
            fg_color='white',
            border_width=1,
            border_color=get_color('Border'),
            corner_radius=10
        )
        frame.pack(fill='x', pady=(0, 15), padx=5)

        inner = ctk.CTkFrame(frame, fg_color='transparent')
        inner.pack(fill='x', padx=20, pady=15)

        # Checkbox
        checkbox = ctk.CTkCheckBox(
            inner,
            text='☑️  Mostrar Filtros Avanzados',
            variable=self.show_filters_var,
            font=('Montserrat', 13, 'bold'),
            command=self.toggle_advanced_filters,
            fg_color=get_color('Sky Blue'),
            hover_color=get_color('Sea Blue'),
            text_color=get_color('Sea Blue'),
            border_width=2,
            border_color=get_color('Border')
        )
        checkbox.pack(side='left')

        # Contenedor padre para insertar los filtros avanzados después
        self.filters_parent = parent

    def toggle_advanced_filters(self):
        """Mostrar/Ocultar filtros avanzados"""
        if self.show_filters_var.get():
            # Mostrar filtros
            if not self.advanced_filters_frame:
                self._create_advanced_filters_content()
            self.advanced_filters_frame.pack(fill='x', pady=(0, 15), padx=5, before=self.filters_parent.winfo_children()[-3])
        else:
            # Ocultar filtros
            if self.advanced_filters_frame:
                self.advanced_filters_frame.pack_forget()

    def _create_advanced_filters_content(self):
        """Crear frame de filtros avanzados"""
        self.advanced_filters_frame = ctk.CTkFrame(
            self.filters_parent,
            fg_color='#F0F8FF',  # Azul muy claro
            border_width=2,
            border_color=get_color('Sky Blue'),
            corner_radius=10
        )

        # Título
        title = ctk.CTkLabel(
            self.advanced_filters_frame,
            text='🔧  Filtros Avanzados',
            font=('Montserrat', 13, 'bold'),
            text_color=get_color('Sea Blue')
        )
        title.pack(pady=(15, 10), padx=20, anchor='w')

        # Container para filtros
        filters_container = ctk.CTkFrame(self.advanced_filters_frame, fg_color='transparent')
        filters_container.pack(fill='x', padx=20, pady=(0, 15))

        # Filtro por División
        label1 = ctk.CTkLabel(
            filters_container,
            text='División:',
            font=('Arial', 11, 'bold'),
            text_color=get_color('Sea Blue')
        )
        label1.grid(row=0, column=0, padx=(0, 10), sticky='w')

        self.division_filter = ctk.CTkOptionMenu(
            filters_container,
            values=['Todas', 'OPERACIONES', 'RRHH', 'TI', 'COMERCIAL', 'FINANZAS'],
            width=150,
            height=35,
            font=('Arial', 10),
            fg_color=get_color('Surface'),
            button_color=get_color('Sky Blue'),
            button_hover_color=get_color('Sea Blue')
        )
        self.division_filter.grid(row=0, column=1, padx=(0, 20))
        self.division_filter.set('Todas')

        # Filtro por Estatus
        label2 = ctk.CTkLabel(
            filters_container,
            text='Estatus:',
            font=('Arial', 11, 'bold'),
            text_color=get_color('Sea Blue')
        )
        label2.grid(row=0, column=2, padx=(0, 10), sticky='w')

        self.status_filter = ctk.CTkOptionMenu(
            filters_container,
            values=['Todos', 'Activo', 'Inactivo'],
            width=150,
            height=35,
            font=('Arial', 10),
            fg_color=get_color('Surface'),
            button_color=get_color('Sky Blue'),
            button_hover_color=get_color('Sea Blue')
        )
        self.status_filter.grid(row=0, column=3, padx=(0, 20))
        self.status_filter.set('Todos')

        # Botón Aplicar Filtros
        btn_apply = ctk.CTkButton(
            filters_container,
            text='✓  Aplicar Filtros',
            width=140,
            height=35,
            font=('Arial', 11, 'bold'),
            fg_color=get_color('Success Green'),
            hover_color='#1E8449',
            text_color='white',
            corner_radius=8,
            command=self.aplicar_filtros_avanzados
        )
        btn_apply.grid(row=0, column=4, padx=(20, 0))

    def _create_quick_queries(self, parent):
        """Sección: Consultas Rápidas"""
        frame = ctk.CTkFrame(
            parent,
            fg_color='white',
            border_width=1,
            border_color=get_color('Border'),
            corner_radius=10
        )
        frame.pack(fill='x', pady=(0, 15), padx=5)

        inner = ctk.CTkFrame(frame, fg_color='transparent')
        inner.pack(fill='x', padx=20, pady=15)

        # Label
        label = ctk.CTkLabel(
            inner,
            text='⚡  Consultas Rápidas:',
            font=('Montserrat', 13, 'bold'),
            text_color=get_color('Sea Blue')
        )
        label.pack(side='left', padx=(0, 20))

        # Botón 1: Todos los Usuarios
        btn1 = ctk.CTkButton(
            inner,
            text='📋  Todos los Usuarios',
            width=180,
            height=40,
            font=('Arial', 11, 'bold'),
            fg_color='#ff9f43',  # Naranja
            hover_color='#ee8630',
            text_color='white',
            corner_radius=8,
            command=self.consultar_todos
        )
        btn1.pack(side='left', padx=(0, 10))

        # Botón 2: Ver Datos de Ejemplo
        btn2 = ctk.CTkButton(
            inner,
            text='💾  Ver Datos de Ejemplo',
            width=200,
            height=40,
            font=('Arial', 11, 'bold'),
            fg_color='#10d876',  # Verde
            hover_color='#0ec46d',
            text_color='white',
            corner_radius=8,
            command=self.mostrar_datos_ejemplo
        )
        btn2.pack(side='left')

    def _create_results_table(self, parent):
        """Sección: Tabla de Resultados con tksheet"""
        # Frame contenedor
        results_frame = ctk.CTkFrame(
            parent,
            fg_color='white',
            border_width=1,
            border_color=get_color('Border'),
            corner_radius=10
        )
        results_frame.pack(fill='both', expand=True, pady=(0, 20), padx=5)

        # Header
        header = ctk.CTkFrame(results_frame, fg_color='transparent')
        header.pack(fill='x', padx=20, pady=(15, 10))

        label = ctk.CTkLabel(
            header,
            text='📊  Resultados',
            font=('Montserrat', 14, 'bold'),
            text_color=get_color('Sea Blue')
        )
        label.pack(side='left')

        # Contador de resultados
        self.results_count_label = ctk.CTkLabel(
            header,
            text='0 registros',
            font=('Arial', 10),
            text_color=get_color('Medium Gray')
        )
        self.results_count_label.pack(side='right', padx=(0, 15))

        # Botón Exportar a Excel
        btn_export = ctk.CTkButton(
            header,
            text='📊 Exportar',
            width=120,
            height=32,
            font=('Arial', 10, 'bold'),
            fg_color='#10d876',  # Verde corporativo
            hover_color='#0ec46d',
            text_color='white',
            corner_radius=6,
            command=self.exportar_a_excel
        )
        btn_export.pack(side='right')

        # Tabla tksheet
        self.table_container = ctk.CTkFrame(
            results_frame,
            fg_color='transparent'
        )
        self.table_container.pack(fill='both', expand=True, padx=20, pady=(0, 20))

        # Crear tksheet
        self._init_tksheet()

    def _init_tksheet(self):
        """Inicializar tabla tksheet con 9 columnas - ESTILO EXCEL CORPORATIVO"""
        try:
            from tksheet import Sheet

            self.sheet = Sheet(
                self.table_container,
                headers=[
                    'User ID', 'Nombre', 'Email', 'Unidad', 'División',
                    'Total Módulos', 'Completados', 'En Progreso', 'Registrados'
                ],

                # ═══ ESTILO EXCEL CORPORATIVO ═══
                # Header: Teal/green corporativo (#00bfa5)
                header_bg='#00bfa5',
                header_fg='white',
                header_font=('Arial', 11, 'bold'),

                # Tabla: Fondo blanco con bordes grises
                table_bg='white',
                table_fg='#2C3E50',  # Texto oscuro para legibilidad
                table_font=('Arial', 10),

                # Índice
                index_bg='#f9f9f9',
                index_fg='#2C3E50',

                # Bordes grises visibles (#d0d0d0)
                outline_color='#d0d0d0',

                # Selección con Sky Blue
                table_selected_cells_bg=get_color('Sky Blue'),
                table_selected_cells_fg='white',

                # Dimensiones
                height=400,
                width=1200
            )

            # Habilitar funcionalidades
            self.sheet.enable_bindings((
                "single_select",
                "row_select",
                "column_width_resize",
                "arrowkeys",
                "right_click_popup_menu",
                "rc_select",
                "copy",
            ))

            self.sheet.pack(fill='both', expand=True)

            # Configurar anchos de columna
            self.sheet.column_width(column=0, width=80)   # User ID
            self.sheet.column_width(column=1, width=200)  # Nombre
            self.sheet.column_width(column=2, width=220)  # Email
            self.sheet.column_width(column=3, width=120)  # Unidad
            self.sheet.column_width(column=4, width=120)  # División
            self.sheet.column_width(column=5, width=110)  # Total
            self.sheet.column_width(column=6, width=110)  # Completados
            self.sheet.column_width(column=7, width=110)  # En Progreso
            self.sheet.column_width(column=8, width=110)  # Registrados

        except ImportError:
            error_label = ctk.CTkLabel(
                self.table_container,
                text='⚠️  La librería tksheet no está instalada.\n\nInstala con: pip install tksheet',
                font=('Arial', 12),
                text_color=get_color('Error Red')
            )
            error_label.pack(pady=50)

    # ═══════════════════════════════════════════════════════
    # MÉTODOS DE CONSULTA (CONECTADOS A BD)
    # ═══════════════════════════════════════════════════════

    def cargar_unidades(self):
        """Carga unidades desde la BD con validación de NULL"""
        if not self.cursor:
            logger.warning("No hay cursor de BD disponible - usando datos de fallback")
            return ['CCI', 'SANCHEZ', 'DENNIS', 'HPMX']

        try:
            query = """
                SELECT DISTINCT NombreUnidad
                FROM Instituto_UnidadDeNegocio
                WHERE NombreUnidad IS NOT NULL
                ORDER BY NombreUnidad
            """
            logger.info("Ejecutando query: cargar_unidades")
            logger.debug(f"SQL: {query}")

            self.cursor.execute(query)
            unidades = [row[0] for row in self.cursor.fetchall()]

            if unidades:
                logger.info(f"Unidades cargadas: {len(unidades)} - {unidades}")
                return unidades
            else:
                logger.warning("No se encontraron unidades en la BD - usando fallback")
                return ['CCI', 'SANCHEZ', 'DENNIS', 'HPMX']

        except Exception as e:
            logger.error(f"Error al cargar unidades: {str(e)}")
            return ['CCI', 'SANCHEZ', 'DENNIS', 'HPMX']

    def buscar_por_id(self):
        """Busca usuario por ID con validación y COALESCE"""
        user_id = self.id_entry.get().strip()
        if not user_id:
            messagebox.showwarning('Atención', 'Ingrese un ID de usuario')
            return

        if not self.cursor:
            messagebox.showerror('Error', 'No hay conexión a la base de datos')
            return

        try:
            query = """
                SELECT
                    u.UserId,
                    COALESCE(u.Nombre, 'Sin nombre') as Nombre,
                    COALESCE(u.Email, 'sin.email@hp.com') as Email,
                    COALESCE(un.NombreUnidad, 'Sin unidad') as NombreUnidad,
                    COALESCE(u.Division, 'Sin división') as Division,
                    COUNT(DISTINCT p.IdModulo) as TotalModulos,
                    COALESCE(SUM(CASE WHEN p.EstatusModuloUsuario = 'Completado' THEN 1 ELSE 0 END), 0) as Completados,
                    COALESCE(SUM(CASE WHEN p.EstatusModuloUsuario = 'En Progreso' THEN 1 ELSE 0 END), 0) as EnProgreso,
                    COALESCE(SUM(CASE WHEN p.EstatusModuloUsuario = 'Registrado' THEN 1 ELSE 0 END), 0) as Registrados
                FROM Instituto_Usuario u
                LEFT JOIN Instituto_UnidadDeNegocio un ON u.IdUnidadDeNegocio = un.IdUnidad
                LEFT JOIN Instituto_ProgresoModulo p ON u.UserId = p.UserId
                WHERE u.UserId = ?
                GROUP BY u.UserId, u.Nombre, u.Email, un.NombreUnidad, u.Division
            """
            logger.info(f"Ejecutando buscar_por_id: {user_id}")
            logger.debug(f"SQL: {query}")

            self.cursor.execute(query, (user_id,))
            results = self.cursor.fetchall()

            if results:
                logger.info(f"Usuario encontrado: {user_id}")
                self.mostrar_resultados(results)
            else:
                logger.warning(f"Usuario no encontrado: {user_id}")
                messagebox.showinfo('No encontrado', f'No se encontró el usuario con ID: {user_id}')

        except Exception as e:
            logger.error(f"Error al buscar usuario {user_id}: {str(e)}")
            messagebox.showerror('Error', f'Error al buscar usuario:\n{str(e)}')

    def consultar_por_unidad(self):
        """Consulta usuarios por unidad de negocio con INNER JOIN"""
        unidad = self.unit_menu.get()

        if not self.cursor:
            messagebox.showerror('Error', 'No hay conexión a la base de datos')
            return

        try:
            query = """
                SELECT
                    u.UserId,
                    COALESCE(u.Nombre, 'Sin nombre') as Nombre,
                    COALESCE(u.Email, 'sin.email@hp.com') as Email,
                    un.NombreUnidad,
                    COALESCE(u.Division, 'Sin división') as Division,
                    COUNT(DISTINCT p.IdModulo) as TotalModulos,
                    COALESCE(SUM(CASE WHEN p.EstatusModuloUsuario = 'Completado' THEN 1 ELSE 0 END), 0) as Completados,
                    COALESCE(SUM(CASE WHEN p.EstatusModuloUsuario = 'En Progreso' THEN 1 ELSE 0 END), 0) as EnProgreso,
                    COALESCE(SUM(CASE WHEN p.EstatusModuloUsuario = 'Registrado' THEN 1 ELSE 0 END), 0) as Registrados
                FROM Instituto_Usuario u
                INNER JOIN Instituto_UnidadDeNegocio un ON u.IdUnidadDeNegocio = un.IdUnidad
                LEFT JOIN Instituto_ProgresoModulo p ON u.UserId = p.UserId
                WHERE un.NombreUnidad = ?
                GROUP BY u.UserId, u.Nombre, u.Email, un.NombreUnidad, u.Division
                ORDER BY u.Nombre
            """
            logger.info(f"Ejecutando consultar_por_unidad: {unidad}")
            logger.debug(f"SQL: {query}")

            self.cursor.execute(query, (unidad,))
            results = self.cursor.fetchall()

            logger.info(f"Resultados encontrados: {len(results)} usuarios en {unidad}")
            self.mostrar_resultados(results)

        except Exception as e:
            logger.error(f"Error al consultar unidad {unidad}: {str(e)}")
            messagebox.showerror('Error', f'Error al consultar unidad:\n{str(e)}')

    def consultar_todos(self):
        """Consulta todos los usuarios con COALESCE y TOP (SQL Server syntax)"""
        if not self.cursor:
            logger.warning("No hay conexión a BD - mostrando datos de ejemplo")
            messagebox.showinfo('Sin conexión', 'No hay conexión a la base de datos\nMostrando datos de ejemplo')
            self.mostrar_datos_ejemplo()
            return

        try:
            query = """
                SELECT TOP 100
                    u.UserId,
                    COALESCE(u.Nombre, 'Sin nombre') as Nombre,
                    COALESCE(u.Email, 'sin.email@hp.com') as Email,
                    COALESCE(un.NombreUnidad, 'Sin unidad') as NombreUnidad,
                    COALESCE(u.Division, 'Sin división') as Division,
                    COUNT(DISTINCT p.IdModulo) as TotalModulos,
                    COALESCE(SUM(CASE WHEN p.EstatusModuloUsuario = 'Completado' THEN 1 ELSE 0 END), 0) as Completados,
                    COALESCE(SUM(CASE WHEN p.EstatusModuloUsuario = 'En Progreso' THEN 1 ELSE 0 END), 0) as EnProgreso,
                    COALESCE(SUM(CASE WHEN p.EstatusModuloUsuario = 'Registrado' THEN 1 ELSE 0 END), 0) as Registrados
                FROM Instituto_Usuario u
                LEFT JOIN Instituto_UnidadDeNegocio un ON u.IdUnidadDeNegocio = un.IdUnidad
                LEFT JOIN Instituto_ProgresoModulo p ON u.UserId = p.UserId
                GROUP BY u.UserId, u.Nombre, u.Email, un.NombreUnidad, u.Division
                ORDER BY u.Nombre
            """
            logger.info("Ejecutando consultar_todos (TOP 100)")
            logger.debug(f"SQL: {query}")

            self.cursor.execute(query)
            results = self.cursor.fetchall()

            logger.info(f"Total de usuarios recuperados: {len(results)}")

            # Estadísticas de debugging
            if results:
                total_completados = sum(row[6] for row in results)
                total_en_progreso = sum(row[7] for row in results)
                total_registrados = sum(row[8] for row in results)
                logger.info(f"Estadísticas - Completados: {total_completados}, En Progreso: {total_en_progreso}, Registrados: {total_registrados}")

            self.mostrar_resultados(results)

        except Exception as e:
            logger.error(f"Error al consultar todos: {str(e)}")
            messagebox.showerror('Error', f'Error al consultar usuarios:\n{str(e)}')

    def mostrar_datos_ejemplo(self):
        """Muestra datos de ejemplo si la BD está vacía"""
        datos_ejemplo = [
            ('T123', 'Juan Pérez', 'juan.perez@hp.com', 'HPMX', 'OPERACIONES', 8, 5, 2, 1),
            ('T456', 'María González', 'maria.gonzalez@hp.com', 'SANCHEZ', 'RRHH', 8, 7, 1, 0),
            ('T789', 'Carlos Ramírez', 'carlos.ramirez@hp.com', 'DENNIS', 'TI', 8, 3, 4, 1),
            ('T234', 'Ana Martínez', 'ana.martinez@hp.com', 'CCI', 'COMERCIAL', 8, 8, 0, 0),
            ('T567', 'Luis Fernández', 'luis.fernandez@hp.com', 'HPMX', 'FINANZAS', 8, 2, 5, 1),
        ]
        self.mostrar_resultados(datos_ejemplo)

    def aplicar_filtros_avanzados(self):
        """Aplica filtros avanzados a la consulta"""
        division = self.division_filter.get()
        status = self.status_filter.get()

        if not self.cursor:
            messagebox.showinfo('Info', 'Requiere conexión a base de datos')
            return

        # Construir query con filtros
        where_clauses = []
        params = []

        if division != 'Todas':
            where_clauses.append("u.Division = ?")
            params.append(division)

        # TODO: Implementar filtro por status cuando se agregue campo Activo a Instituto_Usuario

        where_sql = " AND ".join(where_clauses) if where_clauses else "1=1"

        try:
            self.cursor.execute(f"""
                SELECT
                    u.UserId, u.Nombre, u.Email, un.NombreUnidad, u.Division,
                    COUNT(DISTINCT p.IdModulo) as TotalModulos,
                    SUM(CASE WHEN p.EstatusModuloUsuario = 'Completado' THEN 1 ELSE 0 END) as Completados,
                    SUM(CASE WHEN p.EstatusModuloUsuario = 'En progreso' THEN 1 ELSE 0 END) as EnProgreso,
                    SUM(CASE WHEN p.EstatusModuloUsuario = 'Registrado' THEN 1 ELSE 0 END) as Registrados
                FROM Instituto_Usuario u
                LEFT JOIN Instituto_UnidadDeNegocio un ON u.IdUnidadDeNegocio = un.IdUnidad
                LEFT JOIN Instituto_ProgresoModulo p ON u.UserId = p.UserId
                WHERE {where_sql}
                GROUP BY u.UserId, u.Nombre, u.Email, un.NombreUnidad, u.Division
                ORDER BY u.Nombre
            """, tuple(params))

            results = self.cursor.fetchall()
            self.mostrar_resultados(results)

        except Exception as e:
            messagebox.showerror('Error', f'Error al aplicar filtros: {str(e)}')

    def mostrar_resultados(self, data):
        """Muestra resultados en la tabla tksheet con formato Excel corporativo"""
        if not self.sheet:
            return

        # Limpiar tabla
        self.sheet.set_sheet_data([[]])

        if not data:
            self.results_count_label.configure(text='0 registros')
            self.current_data = []
            messagebox.showinfo('Sin resultados', 'No se encontraron datos')
            return

        # Convertir None a 0 en los datos
        processed_data = []
        for row in data:
            processed_row = list(row)
            for i in range(5, 9):  # Columnas numéricas
                if processed_row[i] is None:
                    processed_row[i] = 0
            processed_data.append(processed_row)

        # Guardar datos para exportar
        self.current_data = processed_data

        # Insertar datos
        self.sheet.set_sheet_data(processed_data)

        # Actualizar contador
        self.results_count_label.configure(text=f'{len(data)} registros')

        # Aplicar formato alternado estilo Excel: blanco (#ffffff) y gris claro (#f9f9f9)
        for i in range(len(processed_data)):
            if i % 2 == 1:  # Filas impares en gris claro
                self.sheet.highlight_rows(
                    i,
                    bg='#f9f9f9',
                    fg='#2C3E50'
                )
            else:  # Filas pares en blanco
                self.sheet.highlight_rows(
                    i,
                    bg='#ffffff',
                    fg='#2C3E50'
                )

    def exportar_a_excel(self):
        """Exporta los datos actuales a un archivo Excel con openpyxl"""
        if not self.current_data:
            messagebox.showwarning('Sin datos', 'No hay datos para exportar.\nRealice una consulta primero.')
            return

        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

            # Crear workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Consulta Usuarios"

            # Headers
            headers = ['User ID', 'Nombre', 'Email', 'Unidad', 'División',
                      'Total Módulos', 'Completados', 'En Progreso', 'Registrados']

            # Estilo header: Teal/green (#00bfa5) con texto blanco
            header_fill = PatternFill(start_color='00bfa5', end_color='00bfa5', fill_type='solid')
            header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
            header_alignment = Alignment(horizontal='center', vertical='center')

            # Bordes grises (#d0d0d0)
            border_side = Side(style='thin', color='d0d0d0')
            border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)

            # Escribir headers
            for col_idx, header in enumerate(headers, start=1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment
                cell.border = border

            # Escribir datos con formato alternado
            for row_idx, row_data in enumerate(self.current_data, start=2):
                # Alternar color de fila: blanco (#ffffff) y gris claro (#f9f9f9)
                if row_idx % 2 == 0:
                    fill = PatternFill(start_color='ffffff', end_color='ffffff', fill_type='solid')
                else:
                    fill = PatternFill(start_color='f9f9f9', end_color='f9f9f9', fill_type='solid')

                for col_idx, value in enumerate(row_data, start=1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.fill = fill
                    cell.font = Font(name='Arial', size=10)
                    cell.alignment = Alignment(horizontal='left' if col_idx <= 5 else 'center', vertical='center')
                    cell.border = border

            # Ajustar anchos de columna
            ws.column_dimensions['A'].width = 12  # User ID
            ws.column_dimensions['B'].width = 25  # Nombre
            ws.column_dimensions['C'].width = 30  # Email
            ws.column_dimensions['D'].width = 15  # Unidad
            ws.column_dimensions['E'].width = 18  # División
            ws.column_dimensions['F'].width = 15  # Total Módulos
            ws.column_dimensions['G'].width = 13  # Completados
            ws.column_dimensions['H'].width = 13  # En Progreso
            ws.column_dimensions['I'].width = 13  # Registrados

            # Altura de filas
            ws.row_dimensions[1].height = 35  # Header
            for row_idx in range(2, len(self.current_data) + 2):
                ws.row_dimensions[row_idx].height = 32  # Filas de datos

            # Guardar con timestamp
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f'consulta_usuarios_{timestamp}.xlsx'

            wb.save(filename)
            logger.info(f"Archivo Excel exportado: {filename}")

            messagebox.showinfo(
                'Exportación exitosa',
                f'Datos exportados correctamente a:\n\n{filename}\n\n' +
                f'Total de registros: {len(self.current_data)}'
            )

        except ImportError:
            messagebox.showerror(
                'Error',
                'La librería openpyxl no está instalada.\n\nInstala con: pip install openpyxl'
            )
        except Exception as e:
            logger.error(f"Error al exportar a Excel: {str(e)}")
            messagebox.showerror('Error', f'Error al exportar a Excel:\n{str(e)}')


# Para testing standalone
if __name__ == '__main__':
    root = ctk.CTk()
    root.title("Test Consultas Hutchison")
    root.geometry("1400x900")

    ctk.set_appearance_mode("light")

    panel = ConsultasHutchison(root)
    panel.pack(fill='both', expand=True)

    root.mainloop()
